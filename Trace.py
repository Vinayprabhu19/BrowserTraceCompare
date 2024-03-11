import json
import os
import ijson 
import openpyxl
from openpyxl.styles import PatternFill


exclude=['https://content.powerapps.com/resource/uci-infra/','fluent-ui','fluentui']
def propagate_ts_to_lower_levels(data, ts=None):
    if isinstance(data, dict):
        if "ts" in data:
            ts = data["ts"]
        elif ts is not None:
            data["ts"] = ts
        for value in data.values():
            propagate_ts_to_lower_levels(value, ts)
    elif isinstance(data, list):
        for item in data:
            propagate_ts_to_lower_levels(item, ts)
            
def filter_records_with_request_method(data):
    filtered_records = []
    if isinstance(data, dict):
        if "requestMethod" in data:
            filtered_records.append(data)
        for value in data.values():
            filtered_records.extend(filter_records_with_request_method(value))
    elif isinstance(data, list):
        for item in data:
            filtered_records.extend(filter_records_with_request_method(item))
    return filtered_records


        

def filter_records_with_function_name(data):
    filtered_records = []

    if isinstance(data, dict):
        if "url" in data:
            filtered_records.append(data)
        for value in data.values():
            filtered_records.extend(filter_records_with_function_name(value))
    elif isinstance(data, list):
        for item in data:
            filtered_records.extend(filter_records_with_function_name(item))
    return filtered_records


def get_urls_in_order(list_of_objects):
    
    urls_in_order = []
    for obj in list_of_objects:
        url = obj.get('url')
        if url :
            flag=True
            for e in exclude:
                if e in url:
                    flag=False
                    break
            if flag==True:
                urls_in_order.append(url)
    return urls_in_order


def url_comparison(array1, array2, output_file):
    # Get the length of the shorter array
    min_len = min(len(array1), len(array2))
    wb = openpyxl.Workbook()
    sheet = wb.active
    
    # Initialize a variable to track the row index in the Excel sheet
    row_index = 1
    
    # Set the width of the columns
    sheet.column_dimensions['A'].width = 50
    sheet.column_dimensions['B'].width = 50
    
    # Define fill colors
    light_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    light_green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    light_blue = PatternFill(start_color="00BFFF", end_color="00BFFF", fill_type="solid")
    
    # Iterate through both arrays
    for i in range(min_len):
        if array1[i] == array2[i]:
            print(f"{i}---{i}")
            
            sheet.cell(row=row_index, column=1, value=array1[i]).fill = light_yellow
            sheet.cell(row=row_index, column=2, value=array2[i]).fill = light_yellow
            row_index += 1
        else:
            print(f"{i}----{i}")
            if array1[i] == "":
                sheet.cell(row=row_index, column=1, value="").fill 
                sheet.cell(row=row_index, column=2, value=array2[i]).fill = light_green
                row_index += 1
            elif array2[i] == "":
                sheet.cell(row=row_index, column=1, value=array1[i]).fill = light_blue
                sheet.cell(row=row_index, column=2, value="").fill
                row_index += 1
                
            else:
                print(f"----{i}")
                sheet.cell(row=row_index, column=1, value=array1[i]).fill = light_blue
                sheet.cell(row=row_index, column=2, value="").fill 
                sheet.cell(row=row_index + 1, column=2, value=array2[i]).fill = light_green
                row_index += 2

    # Check if there are extra elements in array1
    for i in range(min_len, len(array1)):
        print(f"{i}---")
        sheet.cell(row=row_index, column=1, value=array1[i]).fill = light_blue
        row_index += 1

    # Check if there are extra elements in array2
    for i in range(min_len, len(array2)):
        print(f"---{i}")
        sheet.cell(row=row_index, column=2, value=array2[i]).fill = light_green
        row_index += 1

    wb.save(output_file)

def processFile(json_data):

    json_data = sorted(json_data, key=lambda x: x.get('ts', 0))
    propagate_ts_to_lower_levels(json_data)
    records_with_request_method = filter_records_with_request_method(json_data)
    records_with_function_call= filter_records_with_function_name(json_data)
    print(len(records_with_request_method))
    print(len(records_with_function_call))
    
    urls = get_urls_in_order(records_with_function_call)
    #for u in urls:
     #   print(u)
    return [records_with_function_call,records_with_request_method,urls]
    


script_dir = os.path.dirname(__file__)
output_file =  os.path.join(script_dir, r'Traces\comparison.xlsx')


# Construct the relative path to the JSON file
file_path1 =os.path.join(script_dir, r'Traces\T1.json')
file_path2 = os.path.join(script_dir, r'Traces\T2.json')
try:
    
    with open(file_path1, 'r') as file:
        json_data = ijson.items(file,"traceEvents.item")
        items = [o for o in json_data]
        T1=processFile(items)
    with open(file_path2, 'r') as file:
        json_data = ijson.items(file,"traceEvents.item")
        items = [o for o in json_data]
        T2=processFile(items)
    url_comparison(T1[2],T2[2],output_file)
except FileNotFoundError:
    print(f"File '{file_path1}' not found.")
except json.JSONDecodeError:
    print(f"Invalid JSON format in file '{file_path1}'.")